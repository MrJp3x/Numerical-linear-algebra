import openpyxl
import numpy as np




def checkMatrixLength(file):
    matrixLength = 0
    for i in range(1, 1000):
        for j in range(1, 1000):
            if file.cell(row=i, column=j).value is not None:
                matrixLength = (i, j)
            elif file.cell(row=i, column=j).value is None:
                break
    return matrixLength


def createMatrixWithXlElement(Path):
    try:
        book = openpyxl.load_workbook(Path)
        sheet = book.active
        matrixLength = checkMatrixLength(sheet)
        lst = []
        for i in range(1, matrixLength[0] + 1):
            for j in range(1, matrixLength[1] + 1):
                sh33t = sheet.cell(row=i, column=j).value
                if sh33t is None:
                    sh33t = 0
                lst.append(float(sh33t))
        arr = np.array(lst)
        matrix = arr.reshape((matrixLength[0], matrixLength[1]))
        mt = matrix.T
        b = np.array([mt[matrixLength[1] - 1]])
        matrix = np.array(
            mt[0:matrixLength[1] - 1]
        )
        return matrix, b.T, matrixLength

    except FileNotFoundError:
        raise FileNotFoundError('No such file or directory')
    except ValueError:
        raise ValueError('Just enter integer or float')


def writef(matrix):
    book = openpyxl.Workbook()
    sheet = book.active
    m, n = matrix.shape
    for i in range(m):
        for j in range(n):
            sheet.cell(row=i + 1, column=j + 1).value = matrix[i, j]

    book.save('write2cell.xlsx')


# if __name__ == '__main__':
#     arr = np.array(
#         [[1, 2, 3, 4],
#          [5, 4, 8, 6],
#          [4, 5, 2, 7]]
#     )
#     path = str('./sample.xlsx')
#     writef(arr, path)
